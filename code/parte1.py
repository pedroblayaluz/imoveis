#Importando bibliotecas necessarias
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import xlrd

#Criando uma matrix com as cores de 'price-info.xlsx'
wb = load_workbook('data/price-info.xlsx')
sheet = wb['Sheet1']
col_array = []
for line in sheet.iter_rows():
  row_array = []
  for c in line:
      row_array.append(c.fill.bgColor.rgb)
  col_array.append(row_array)
  
#Juntando as informaçoes de cores e precos
#DataFrame com os precos, IDs e datas
prices_df = pd.read_excel('data/price-info.xlsx')
#DataFrame com as cores e incluir IDs e datas do outro DataFrame
color_df = pd.DataFrame(np.array(col_array),columns=prices_df.columns)
color_df = color_df.iloc[1:,1:].reset_index(drop=True)
color_df = pd.concat([prices_df.iloc[:,0], color_df], axis=1)
#Unir os dois DataFrames
melt_color = pd.melt(color_df, id_vars='Unnamed: 0', var_name='date', value_name='color')
melt_prices = pd.melt(prices_df, id_vars='Unnamed: 0', var_name='date', value_name='price')
dados_totais = pd.concat([melt_color,melt_prices], axis=1)
dados_totais = dados_totais.loc[:,~dados_totais.columns.duplicated()]
#Criar a variavel ocupado a partir da cor (verde=1, outro=0)
dummies = pd.get_dummies(dados_totais.color)
dados_totais = pd.concat([dados_totais,dummies.FF33CCCC], axis=1)
dados_totais.columns = ['id', 'data', 'cor', 'preco', 'ocupado']
#Criar a variavel faturamento multiplicando ocupado e preco
dados_totais['faturamento'] = dados_totais.preco * dados_totais.ocupado

#Criar a planilha contendo a lista de IDs 
dados_totais['mes'] = pd.DatetimeIndex(dados_totais['data']).month
faturamento_mensal = pd.pivot_table(dados_totais, index=["id"], columns=["mes"], values=["faturamento"], aggfunc=np.sum)
faturamento_mensal.columns = ['jan', 'feb', 'jun', 'jul', 'ago','set','out','nov','dez']
faturamento_mensal = faturamento_mensal[['jun','jul','ago','set','out','nov','dez','jan']] #Reordenando e tirando fevereiro por ter poucos dados

#Removendo fevereiro de dados_totais
dados_totais = dados_totais.loc[dados_totais['mes'] != 2]
#Salvando arquivo
faturamento_mensal.to_csv('data/faturamento_mensal.csv')
dados_totais.to_csv('data/dados_totais.csv')
